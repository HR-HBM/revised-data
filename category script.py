import random

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

dataframe = openpyxl.load_workbook("sorted links.xlsx")
print("workbook loaded")

dataframe1 = dataframe.active
links_list = []
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1):
        links = col[row].value
        links_list.append(links)

# proxies = ['http://72.10.160.93:15099', 'http://72.10.164.178:2721', 'http://172.93.213.177:80',
#            'socks4://47.250.177.202:3128', 'http://72.10.160.90:26279', 'http://50.62.183.223:80',
#            'http://67.43.228.253:11435', 'socks4://98.170.57.249:4145', 'socks4://109.236.81.77:11028',
#            'http://152.26.231.77:9443', 'http://72.10.164.178:9631', 'socks4://174.77.111.196:4145',
#            'http://218.1.142.120:57114', 'http://67.43.227.226:27901', 'http://152.26.229.93:9443',
#            'http://72.10.164.178:30273', 'http://47.250.11.111:8081', 'http://153.101.67.170:9002',
#            'http://202.61.206.250:8888', 'http://67.43.228.250:25981', 'http://47.250.11.111:80',
#            'http://183.234.215.11:8443', 'http://195.234.62.51:80', 'http://65.108.195.47:8080',
#            'http://209.141.58.24:80', 'http://118.117.189.242:8089', 'http://58.246.58.150:9002',
#            'http://67.43.236.20:20041', 'http://152.26.229.34:9443', 'http://72.10.160.90:2819',
#            'http://154.203.132.55:8090', 'http://67.43.227.227:3217', 'http://89.145.162.81:3128',
#            'http://160.86.242.23:8080', 'http://72.10.164.178:23015', 'http://72.10.160.173:8597',
#            'http://72.10.164.178:18445', 'socks4://137.184.182.145:9338', 'http://72.10.160.90:16575',
#            'http://72.10.164.178:26533', 'socks4://31.135.91.9:4145', 'http://8.211.42.167:9098',
#            'http://72.10.164.178:25363', 'http://72.10.160.172:23075', 'http://171.88.121.171:8118',
#            'http://152.26.229.46:9443', 'socks4://192.111.130.5:17002', 'socks4://125.109.59.252:8081',
#            'http://183.162.197.197:8060', 'http://72.10.160.91:22119', 'http://154.65.39.7:80',
#            'socks4://103.60.187.1:52195', 'http://120.234.203.171:9002', 'http://67.43.227.227:1869',
#            'http://141.145.214.176:80', 'http://47.238.128.246:4145', 'socks4://88.80.150.102:4435',
#            'socks4://188.187.0.158:8080', 'http://67.43.227.227:1063', 'http://72.10.160.93:28905',
#            'http://67.43.236.20:6221', 'http://211.104.20.205:8080', 'http://5.161.115.29:51111',
#            'http://42.3.187.52:8080', 'socks4://197.232.36.85:41890', 'http://220.248.70.237:9002',
#            'http://118.117.189.196:8089', 'http://159.65.237.225:1951', 'socks4://142.54.235.9:4145',
#            'http://173.249.30.197:8118', 'socks4://47.74.46.81:8081', 'socks4://67.43.227.230:10176',
#            'socks4://118.179.195.20:4145', 'socks4://86.42.188.17:8081', 'http://116.111.115.33:10013',
#            'http://72.10.164.178:25141', 'socks4://200.195.174.146:3128', 'socks4://23.97.56.11:3128',
#            'http://43.134.33.254:3128', 'http://152.26.229.86:9443', 'http://67.43.228.252:19227',
#            'http://72.10.164.178:1995', 'socks4://205.178.136.41:8447', 'http://122.152.4.135:6000',
#            'http://52.82.123.144:3128', 'http://72.10.160.90:29891', 'http://67.43.227.227:29645',
#            'http://125.77.25.178:8090', 'http://67.43.228.253:14051', 'http://72.10.160.173:6181',
#            'http://72.10.160.174:12283', 'http://162.19.107.209:3128', 'http://47.252.29.28:11222',
#            'http://72.10.164.178:6557', 'http://110.232.67.43:55443', 'socks4://142.54.228.193:4145',
#            'http://72.10.160.173:7389', 'http://111.238.48.221:80', 'http://72.10.164.178:20611',
#            'http://72.10.160.173:11105', 'http://67.43.227.228:27623', 'http://109.236.83.153:8888',
#            'http://67.43.227.227:25933', 'http://152.26.229.83:9443', 'http://188.116.22.67:24099',
#            'http://171.234.237.179:10004', 'socks4://124.41.213.174:5678', 'socks4://172.105.107.219:12334',
#            'http://139.213.197.139:8118', 'http://146.59.202.70:80', 'http://67.43.228.250:3509',
#            'http://67.43.228.253:1557', 'http://51.89.255.67:80', 'http://180.88.111.187:3128',
#            'http://67.43.227.227:17447', 'http://159.65.237.225:1822', 'http://72.10.164.178:21641',
#            'http://43.255.113.232:8080', 'socks4://75.162.131.135:3128', 'http://47.91.120.190:9080',
#            'http://67.43.236.20:2705', 'http://135.181.154.225:80', 'http://190.103.177.131:80',
#            'http://72.10.160.170:30373', 'http://80.249.112.162:80', 'socks4://24.249.199.12:4145',
#            'http://185.105.90.88:4444', 'http://35.94.217.206:3128', 'http://104.129.194.99:8800',
#            'http://61.129.2.212:8080', 'http://47.91.120.190:8080', 'http://72.10.160.90:5231',
#            'http://72.10.160.94:19283', 'http://72.10.164.178:27879', 'http://72.10.160.172:2633',
#            'http://27.112.70.42:8083', 'http://72.10.160.170:1763', 'http://14.23.152.222:9090',
#            'http://185.105.88.63:4444', 'http://72.10.164.178:20777', 'http://121.234.73.237:9002',
#            'http://125.77.25.178:8080', 'socks4://110.74.195.2:4153', 'http://67.43.236.20:30859',
#            'socks4://47.76.144.139:8443', 'http://128.199.202.122:3128', 'socks4://154.12.253.232:64196',
#            'http://67.43.227.227:1119', 'http://14.199.30.127:80', 'http://31.128.43.98:8118',
#            'http://67.43.228.252:32147', 'http://49.70.172.179:8089', 'http://72.10.160.170:14099',
#            'http://164.163.185.199:80', 'http://95.216.215.36:80', 'http://8.213.151.128:3128',
#            'http://72.10.160.171:14269', 'socks4://184.185.2.12:4145', 'socks4://104.236.0.129:50503',
#            'http://72.10.164.178:27349', 'http://198.49.68.80:80', 'http://34.135.166.24:80',
#            'http://67.43.227.226:3987', 'socks4://103.97.94.22:4153', 'http://117.6.55.135:10004',
#            'http://67.43.227.227:20627', 'http://67.43.236.18:27043', 'http://72.10.160.174:13667',
#            'http://43.134.121.40:3128', 'http://72.10.160.170:25129', 'http://51.89.14.70:80', 'http://85.215.64.49:80',
#            'http://111.59.4.88:9002', 'http://47.242.47.64:8888', 'socks4://185.215.53.217:3629',
#            'http://165.232.129.150:80', 'http://74.103.66.15:80', 'http://72.10.164.178:21807',
#            'http://67.43.236.20:18293', 'http://116.111.116.106:10004', 'socks4://46.109.146.244:4145',
#            'socks4://104.207.36.58:3128', 'socks4://120.138.22.107:37616', 'http://117.6.55.135:10006',
#            'http://72.10.160.94:10239', 'http://72.10.160.90:23627', 'http://162.223.116.54:80',
#            'http://72.10.160.173:29019', 'http://111.224.216.59:8089', 'http://72.10.160.172:18665',
#            'http://217.64.149.146:8118', 'socks4://149.11.44.58:10931', 'http://67.43.227.228:5435',
#            'socks4://103.115.255.225:36331', 'http://139.59.1.14:3128', 'socks4://79.103.85.90:8080',
#            'socks4://180.234.87.113:8080', 'socks4://117.74.65.207:10006', 'socks4://122.3.139.85:1080',
#            'http://162.214.165.203:80', 'http://67.43.228.253:29835', 'socks4://116.118.98.25:5678',
#            'socks4://103.189.234.161:1080', 'socks4://166.0.235.2:61206', 'socks4://199.229.254.129:4145',
#            'http://194.182.187.78:3128', 'http://72.10.164.178:6845', 'socks4://8.130.90.177:8008',
#            'socks4://44.226.95.55:3128', 'socks4://138.186.222.129:5678', 'http://72.10.160.173:12283',
#            'http://84.252.75.136:4444', 'http://72.10.160.90:23881', 'socks4://38.133.200.94:31596',
#            'http://183.215.23.242:9091', 'http://72.10.160.170:2563', 'http://159.65.237.225:1637',
#            'socks4://45.43.82.153:6147', 'socks4://75.110.125.218:80', 'http://199.168.175.179:80',
#            'http://8.212.151.166:8080', 'http://47.237.113.119:1720', 'http://72.10.160.90:11097',
#            'http://67.43.228.253:21055', 'http://116.63.129.202:6000', 'http://67.43.236.18:5579',
#            'http://91.65.103.3:80', 'http://185.105.89.249:4444', 'http://38.65.81.118:8080',
#            'socks4://124.90.45.58:10800', 'http://111.224.212.210:8089', 'http://72.10.160.91:19283',
#            'socks4://8.130.54.67:8080', 'socks4://108.181.133.58:22767', 'http://72.10.164.178:22313',
#            'http://72.10.160.170:24813', 'socks4://89.237.34.193:51549', 'http://152.26.231.83:9443',
#            'socks4://85.206.167.138:31951', 'http://36.134.91.82:8888', 'socks4://212.83.142.145:64494',
#            'http://72.10.164.178:27527', 'http://67.43.227.230:10125', 'http://118.117.189.189:8089',
#            'http://154.73.111.133:1981', 'socks4://190.2.110.7:4153', 'http://159.54.145.18:80',
#            'http://67.43.227.227:28143', 'http://67.43.228.250:24549', 'socks4://45.120.38.246:1088',
#            'http://116.111.123.96:10005', 'http://64.227.133.223:9090', 'http://67.43.228.250:10899',
#            'socks4://181.57.194.28:5678', 'http://67.43.227.228:2863', 'socks4://67.11.89.131:8888',
#            'socks4://49.229.36.170:4153', 'socks4://5.39.69.35:56650', 'socks4://110.78.186.70:8080',
#            'http://67.43.227.227:30667', 'http://220.73.144.83:10808', 'http://148.72.140.24:30117',
#            'http://190.122.185.170:999', 'socks4://114.80.130.220:10884', 'http://101.101.217.36:80',
#            'socks4://156.235.51.186:6500', 'http://45.92.177.60:8080', 'http://72.10.164.178:7461',
#            'http://84.252.73.132:4444', 'http://67.43.227.228:12123', 'http://185.217.198.121:4444',
#            'socks4://46.8.22.133:5500', 'http://152.26.229.42:9443', 'http://152.32.173.226:8199',
#            'socks4://77.104.227.62:80', 'http://8.209.255.13:3128', 'http://152.26.231.86:9443',
#            'http://185.217.199.176:4444', 'http://152.26.231.22:9443', 'http://152.26.229.57:9443',
#            'http://67.43.227.227:24777', 'http://103.162.50.13:80', 'socks4://47.76.144.139:8080',
#            'socks4://138.201.21.228:62581', 'http://116.111.116.106:10010', 'http://49.70.172.146:8089',
#            'http://67.43.236.19:7125', 'http://43.255.113.232:84', 'http://67.43.227.226:25407',
#            'http://154.64.226.138:80', 'http://45.9.75.76:4444', 'http://67.43.228.250:7117',
#            'http://116.111.123.96:10004', 'socks4://185.230.15.241:3128', 'http://103.141.180.254:8080',
#            'http://122.10.225.55:8000', 'http://72.10.160.90:26201', 'http://72.10.160.171:11021',
#            'http://67.43.228.250:17283', 'http://72.10.160.90:11513', 'http://159.54.149.67:80',
#            'http://118.117.189.222:8089', 'http://222.89.237.101:9002', 'socks4://103.25.45.38:4153',
#            'http://4.234.77.255:8080', 'http://64.92.82.59:8080', 'http://72.10.164.178:12145',
#            'http://152.26.229.47:9443', 'http://72.10.160.170:30551', 'http://72.10.160.170:25273',
#            'http://67.43.227.227:1627', 'http://116.203.28.43:80', 'http://45.22.209.157:8888',
#            'http://67.43.236.20:32449', 'socks4://141.226.246.176:10171', 'http://47.237.2.245:1311',
#            'http://36.37.86.26:9812', 'http://66.29.154.103:3128', 'http://27.147.155.42:58080',
#            'socks4://42.41.208.95:8888', 'socks4://43.159.28.112:19739', 'http://67.43.228.253:16843',
#            'socks4://67.213.212.53:30372', 'http://185.105.91.62:4444', 'http://203.154.162.230:443',
#            'http://190.152.5.17:39888', 'http://67.43.228.250:17029', 'http://67.43.228.253:11041',
#            'socks4://1.172.173.93:8088', 'http://67.43.227.227:3987', 'http://152.26.231.42:9443',
#            'socks4://160.251.72.223:29532', 'socks4://67.213.212.58:41429', 'socks4://17.57.227.222:1524',
#            'socks4://103.109.244.4:10667', 'http://67.43.227.227:9045', 'http://67.43.228.250:8711',
#            'http://72.10.164.178:19643', 'socks4://205.150.131.39:3306', 'socks4://174.77.111.197:4145',
#            'socks4://116.99.239.65:30062', 'http://185.232.169.108:4444', 'socks4://212.83.142.114:33009',
#            'http://162.223.90.130:80', 'http://72.10.164.178:30429', 'http://4.155.2.13:9480',
#            'http://35.197.150.32:8888', 'http://46.23.148.147:49868', 'http://51.75.206.209:80',
#            'http://72.10.160.170:2633', 'socks4://222.165.223.138:41541', 'socks4://198.0.142.181:4122',
#            'http://67.43.227.227:11265', 'http://67.43.227.226:23341', 'socks4://107.180.93.248:46338',
#            'http://209.126.80.197:9595', 'http://131.153.187.5:51410', 'http://122.185.198.242:7999',
#            'http://5.58.33.187:55507', 'http://133.18.234.13:80', 'http://208.163.39.218:53281',
#            'socks4://64.124.145.1:1080', 'socks4://92.205.110.118:10147', 'http://193.122.61.167:80',
#            'http://154.65.39.8:80', 'socks4://163.172.166.35:16379', 'socks4://38.62.223.163:3128',
#            'socks4://39.104.23.154:8080', 'http://66.206.15.148:8136', 'socks4://212.83.143.97:47311',
#            'socks4://163.5.71.220:3128', 'socks4://117.74.65.207:80', 'http://38.54.6.39:80', 'http://4.155.2.13:9443',
#            'http://67.43.227.227:32701', 'socks4://43.229.11.206:5844', 'http://72.10.160.92:23233',
#            'http://72.10.164.178:15341', 'http://72.10.164.178:4763', 'http://72.10.160.170:12283',
#            'http://67.43.227.227:26895', 'http://152.26.231.93:9443', 'http://58.20.248.139:9002',
#            'http://67.43.227.227:10471', 'socks4://44.226.52.95:8080', 'http://72.10.160.171:14491',
#            'socks4://102.254.41.197:1337', 'http://67.43.228.250:28763', 'socks4://31.28.5.185:63909',
#            'http://67.43.236.20:27043', 'http://72.10.160.90:17813', 'http://195.66.93.188:3128',
#            'http://72.10.160.170:3483', 'socks4://92.241.112.206:1080', 'http://72.10.160.90:22371',
#            'http://152.26.231.94:9443', 'http://72.10.160.170:14927', 'http://43.255.113.232:86',
#            'http://80.80.163.190:46276', 'http://72.10.164.178:19635', 'http://72.10.160.172:32115',
#            'http://67.43.236.20:13693', 'http://67.43.236.20:1499', 'http://72.10.160.90:20467',
#            'http://67.43.236.20:2671', 'http://123.126.158.50:80', 'http://124.248.180.163:8080',
#            'http://152.32.176.241:8443', 'http://67.43.227.229:10125', 'http://23.247.136.245:80',
#            'http://84.252.74.190:4444', 'http://67.43.228.250:27437', 'http://72.10.160.171:10203',
#            'http://67.43.236.19:20203', 'http://67.43.227.226:26553', 'http://72.10.160.171:9571',
#            'http://72.10.164.178:27159', 'http://35.209.198.222:80', 'http://103.148.45.60:8833',
#            'http://72.10.164.178:4687', 'http://72.10.160.171:21419', 'http://43.255.113.232:83',
#            'http://67.43.227.228:28143', 'http://67.43.236.20:12241', 'http://62.72.56.132:80',
#            'http://72.10.164.178:19289', 'http://5.75.142.128:3128', 'http://91.229.118.126:3128',
#            'socks4://212.83.137.94:13912', 'socks4://122.228.92.103:80', 'socks4://67.43.228.253:25939',
#            'socks4://27.147.140.5:8080', 'socks4://67.43.227.227:2073', 'socks4://201.216.239.162:1080',
#            'http://20.106.146.212:5003', 'socks4://159.255.182.200:8080', 'socks4://193.216.224.108:8192',
#            'socks4://66.29.129.56:16195', 'socks4://38.56.23.33:999', 'socks4://5.78.83.35:8080',
#            'socks4://163.204.241.136:9999', 'socks4://95.216.29.208:13079', 'socks4://148.72.212.198:55587',
#            'socks4://86.71.12.89:1080', 'socks4://185.122.204.56:31125', 'socks4://113.160.154.7:4153',
#            'socks4://168.232.60.62:5678', 'socks4://108.181.132.115:24979', 'socks4://49.0.252.39:1337',
#            'socks4://8.130.39.117:3128', 'socks4://38.156.73.36:8080', 'socks4://212.83.142.100:41897',
#            'socks4://202.136.88.89:8080', 'socks4://45.61.127.68:6007', 'socks4://103.212.128.172:39149',
#            'socks4://167.172.159.43:40830', 'socks4://105.214.21.72:5678', 'socks4://47.91.115.179:80',
#            'socks4://156.237.0.239:6500', 'socks4://67.205.158.205:80', 'socks4://148.72.215.79:28509',
#            'socks4://37.44.238.2:60796', 'socks4://103.205.135.225:6969', 'socks4://5.61.205.217:4153',
#            'socks4://36.50.135.76:27434', 'socks4://123.56.173.64:80', 'socks4://198.100.156.43:42411',
#            'socks4://67.43.236.18:2959', 'socks4://192.111.139.162:4145', 'socks4://8.213.137.155:8060',
#            'socks4://186.219.96.47:49923', 'socks4://103.68.0.242:5678', 'socks4://67.213.210.167:52717',
#            'socks4://38.54.101.254:80', 'socks4://3.79.103.57:8080', 'socks4://38.54.95.19:3128',
#            'socks4://210.45.200.5:11579', 'socks4://8.130.71.75:8081', 'socks4://72.167.220.46:16132',
#            'socks4://60.146.78.9:3128', 'socks4://67.43.227.227:32335', 'socks4://8.243.169.14:8080',
#            'socks4://205.178.136.80:8447', 'socks4://39.102.211.162:8080', 'socks4://65.21.232.59:8786',
#            'socks4://47.238.128.246:8443', 'socks4://217.150.77.31:53281', 'socks4://167.103.19.22:11194',
#            'http://116.103.130.17:26807', 'http://116.103.130.17:26802', 'socks4://206.42.40.0:5678',
#            'socks4://162.0.220.220:11600', 'socks4://179.27.86.36:4153', 'socks4://203.96.177.211:13461',
#            'socks4://146.120.160.148:5678', 'socks4://194.146.110.211:8888', 'socks4://103.60.138.65:4153',
#            'socks4://107.6.181.221:3128', 'socks4://35.77.195.189:58308', 'socks4://212.170.60.124:4145',
#            'socks4://82.157.172.60:1080', 'socks4://164.52.42.2:4145', 'socks4://31.207.47.29:13380',
#            'socks4://223.84.6.245:8123', 'socks4://82.110.30.53:8081', 'socks4://103.147.246.131:1080',
#            'socks4://202.124.46.100:4145', 'socks4://159.223.166.21:5199', 'socks4://45.61.188.134:44499',
#            'socks4://157.20.63.39:31324', 'socks4://198.23.239.221:6627', 'socks4://45.134.79.45:10520',
#            'socks4://14.102.152.10:3125', 'socks4://181.209.82.154:14888', 'socks4://20.229.131.220:1080',
#            'socks4://122.121.225.176:80', 'socks4://120.46.215.52:4153', 'socks4://147.5.17.173:47657',
#            'socks4://185.134.99.61:4153', 'socks4://162.245.239.110:63947', 'socks4://195.2.84.34:32365',
#            'socks4://47.251.87.74:9080', 'socks4://107.180.90.88:32168', 'socks4://216.173.120.230:6522',
#            'socks4://90.188.56.158:4153', 'socks4://243.77.45.179:80', 'http://192.140.42.83:31511',
#            'socks4://160.251.74.65:18202', 'socks4://195.177.217.131:1270', 'socks4://67.205.177.122:30375',
#            'socks4://89.39.107.216:11478', 'socks4://72.10.164.178:19911', 'socks4://184.168.121.153:51206',
#            'socks4://173.212.237.43:35781', 'socks4://177.106.0.129:3128', 'socks4://67.43.228.253:31703',
#            'socks4://149.129.135.16:6666', 'socks4://212.83.142.149:11533', 'socks4://49.48.42.15:8080',
#            'socks4://103.21.244.213:80', 'socks4://104.207.46.249:3128', 'socks4://36.95.231.205:5678',
#            'socks4://177.131.29.213:4153', 'socks4://102.244.120.10:45413', 'socks4://51.222.241.157:39617',
#            'socks4://199.168.138.123:8080', 'socks4://64.185.0.17:8080', 'socks4://202.162.212.163:4153',
#            'socks4://167.114.82.169:8080', 'socks4://152.42.229.153:3128', 'socks4://154.79.246.18:9898',
#            'socks4://47.91.95.174:9002', 'socks4://41.184.212.3:4153', 'socks4://37.221.162.24:3128',
#            'socks4://103.123.231.202:8080', 'socks4://67.43.236.20:10579', 'socks4://45.63.39.16:20118',
#            'http://122.160.30.99:80', 'socks4://222.124.204.66:8080', 'socks4://193.230.242.29:8080',
#            'socks4://164.92.237.188:62586', 'socks4://186.211.2.54:4145', 'socks4://179.108.181.73:4153',
#            'socks4://186.96.124.242:4153', 'socks4://67.213.212.40:45974', 'http://58.220.95.66:11143',
#            'socks4://179.125.172.177:4153', 'socks4://61.135.155.82:443', 'socks4://167.172.123.221:9200',
#            'socks4://174.75.211.222:4145', 'socks4://122.98.218.104:443', 'socks4://67.213.212.54:40747',
#            'socks4://134.35.153.207:1080', 'socks4://38.91.106.214:13623', 'socks4://67.43.227.226:15023',
#            'socks4://135.201.62.207:1337', 'socks4://118.71.99.233:5678', 'socks4://212.83.143.151:22423',
#            'socks4://67.201.59.70:4145', 'socks4://8.5.77.234:3128', 'socks4://150.140.193.137:3128',
#            'socks4://213.238.168.109:80', 'socks4://175.171.109.165:8081', 'socks4://185.226.113.180:38030',
#            'socks4://172.105.21.57:1967', 'socks4://8.211.49.86:8008', 'socks4://190.210.62.131:6473',
#            'socks4://190.109.72.17:33633', 'socks4://117.171.233.178:8123', 'socks4://143.255.140.28:5678',
#            'socks4://171.221.174.230:10800', 'socks4://104.200.16.220:61922', 'socks4://164.92.86.113:52494',
#            'socks4://66.29.129.53:61693', 'socks4://45.116.114.37:5678', 'socks4://115.175.212.45:53',
#            'socks4://72.10.164.178:28991', 'socks4://172.99.151.3:3389', 'socks4://162.55.95.91:20128',
#            'http://103.253.103.50:80', 'socks4://67.213.212.53:51004', 'socks4://200.26.26.34:3128',
#            'socks4://196.3.97.82:23500', 'socks4://117.102.115.154:4153', 'socks4://177.71.229.43:8888',
#            'socks4://198.12.253.239:63554', 'socks4://84.26.227.165:1337', 'socks4://50.206.25.111:80',
#            'socks4://162.19.7.47:63029', 'socks4://70.166.167.38:57728', 'socks4://185.32.4.65:4153',
#            'socks4://213.166.140.6:8080', 'socks5://113.121.248.200:8118', 'socks4://72.214.108.67:4145',
#            'socks4://8.215.15.163:80', 'socks4://212.126.5.246:42344', 'socks4://200.24.159.191:999',
#            'socks4://162.0.220.214:46895', 'socks4://91.229.23.206:35965', 'socks4://47.90.149.238:80',
#            'http://87.248.129.26:80', 'socks4://100.1.53.24:5678', 'socks4://67.213.212.39:32259',
#            'socks4://8.213.128.90:9000', 'socks4://185.161.186.92:54321', 'socks4://102.69.176.98:10081',
#            'socks4://96.36.50.99:39593', 'http://104.129.192.170:10878', 'socks4://47.250.135.180:10080',
#            'http://133.232.81.141:80', 'socks4://117.5.22.234:5308', 'socks4://37.52.13.164:5678',
#            'socks4://108.181.132.115:48083', 'socks4://107.180.101.18:37486', 'socks4://190.138.250.48:3629',
#            'socks4://93.190.143.87:37763', 'socks4://139.180.209.22:15202', 'socks4://117.206.155.82:34883',
#            'socks4://172.105.21.57:2224', 'socks4://116.99.228.229:24434', 'socks4://162.144.103.99:60035',
#            'socks4://72.195.34.42:4145', 'socks4://34.125.122.252:8585', 'socks4://138.201.21.232:46617',
#            'socks4://166.0.235.142:62498', 'socks4://78.128.124.108:5678', 'socks4://84.93.224.76:8888',
#            'socks4://149.129.255.179:8443', 'socks4://136.226.245.22:8080', 'socks4://148.72.212.125:46427',
#            'socks4://67.43.236.20:1625', 'socks4://86.111.144.10:4145', 'http://38.54.6.39:312',
#            'socks4://159.89.169.49:40598']
#

# def get_random_proxy():
#     return random.choice(proxies)


def get_random_user_agent():
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36",
        "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:89.0) Gecko/20100101 Firefox/89.0",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/47.0.2526.111 Safari/537.36",
        "Mozilla/5.0 (Linux; Android 10; Pixel 3 XL Build/QP1A.190711.020) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Mobile Safari/537.36",
        "Mozilla/5.0 (iPhone; CPU iPhone OS 14_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15E148 Safari/604.1",
        "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3",
        "Mozilla/5.0 (iPad; CPU OS 14_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15E148 Safari/604.1",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:91.0) Gecko/20100101 Firefox/91.0",
        "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:60.0) Gecko/20100101 Firefox/60.0"
    ]
    return random.choice(user_agents)


chrome_options = webdriver.ChromeOptions()
chrome_options.add_experimental_option("detach", True)

user_agent = get_random_user_agent()
chrome_options.add_argument(f"user-agent={user_agent}")

# # proxy = get_random_proxy()
# chrome_options.add_argument('--proxy-server=http://72.10.160.90:26279')

driver = webdriver.Chrome(options=chrome_options)

maxLink = dataframe1.max_row + 1
currentLinkIndex = 5614


while currentLinkIndex <= maxLink:

    print(links_list[currentLinkIndex])
    driver.get(f"{links_list[currentLinkIndex]}")


    # wait = WebDriverWait(driver, 10)
    # item_data = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='finder-profile']/div/div/section/div/div/div[2]/div[5]/button/div/span")))
    #
    # driver.execute_script("arguments[0].click();", item_data)

    time.sleep(2)

    # name_text = ""
    company_text = ""
    category_text = []
    formatted_categories = ""
    # number_text = ""
    # business_area_text = ""
    # website_text = ""

    try:
        company = driver.find_element(By.XPATH, value="//*[@id='finder-profile']/div/div/section/div/div/div[1]/h1")
        company_text = company.text
    except:
        pass

    try:
        html_list = driver.find_element(By.XPATH, value="//*[@id='finder-products']/div/div/section/div[2]/div/div/div[1]/ul")
        category = html_list.find_elements(By.TAG_NAME, value="li")
        for item in category:
            category_text.append(item.text)
            formatted_categories = ', '.join(category_text)
            print(formatted_categories)
        # category = driver.find_element(By.XPATH, value="//*[@id='profile-title']/h1")
        # name_text = category.text
    except:
        pass



    # try:
    #     email = driver.find_element(By.XPATH, value="//*[@id='profile-business-data']/div[1]/div[2]/div[1]/div[1]")
    #     email_text = email.text
    # except:
    #     pass
    #
    # try:
    #     number = driver.find_element(By.XPATH, value="//*[@id='profile-business-data']/div[1]/div[2]/div[1]/div[2]")
    #     number_text = number.text
    # except:
    #     pass
    #
    # try:
    #     business_area = driver.find_element(By.XPATH,
    #                                         value="//*[@id='profile-business-data']/div[2]/div/table/tbody/tr[3]/td[2]")
    #     business_area_text = business_area.text
    # except:
    #     pass
    #
    # try:
    #     website = driver.find_element(By.XPATH,
    #                                   value="//*[@id='profile-business-data']/div[1]/div[2]/div[1]/div[3]/div[2]/ul/li/a")
    #     website_text = website.text
    # except:
    #     pass

    driver.execute_script("window.open('https://forms.gle/jBGnYSegBUZoDANT9', '_blank');")
    wait = WebDriverWait(driver, 10)


    print("form accessed")
    # Switch to the Google Form tab
    driver.switch_to.window(driver.window_handles[-1])
    time.sleep(3)

    print("form loaded")

    company_name = wait.until(EC.presence_of_element_located((By.XPATH,"//*[@id='mG61Hd']/div[2]/div/div[2]/div[1]/div/div/div[2]/div/div[1]/div[2]/textarea")))

    # company_name = driver.find_element(By.XPATH,value="//*[@id='mG61Hd']/div[2]/div/div[2]/div[1]/div/div/div[2]/div/div[1]/div[2]/textarea")
    company_name.send_keys(company_text)

    category_name = driver.find_element(By.XPATH, value="//*[@id='mG61Hd']/div[2]/div/div[2]/div[2]/div/div/div[2]/div/div[1]/div[2]/textarea")
    category_name.send_keys(formatted_categories)

    # company_category = driver.find_element(By.XPATH,value="//*[@id='mG61Hd']/div[2]/div/div[2]/div[3]/div/div/div[2]/div/div[1]/div[2]/textarea")
    # company_category.send_keys(business_area_text)
    #
    # company_number = driver.find_element(By.XPATH,value="//*[@id='mG61Hd']/div[2]/div/div[2]/div[4]/div/div/div[2]/div/div[1]/div/div[1]/input")
    # company_number.send_keys(number_text)
    #
    # company_email = driver.find_element(By.XPATH,value="//*[@id='mG61Hd']/div[2]/div/div[2]/div[5]/div/div/div[2]/div/div[1]/div/div[1]/input")
    # company_email.send_keys(email_text)
    #
    # company_website = driver.find_element(By.XPATH,value="//*[@id='mG61Hd']/div[2]/div/div[2]/div[6]/div/div/div[2]/div/div[1]/div/div[1]/input")
    # company_website.send_keys(website_text)

    submit = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='mG61Hd']/div[2]/div/div[3]/div[1]/div[1]/div")))

    submit.send_keys(Keys.ENTER)

    driver.close()

    driver.switch_to.window(driver.window_handles[-1])
    print(currentLinkIndex)

    currentLinkIndex += 1
    # index_limit = [5, 10, 15, 20, 25]
    if currentLinkIndex % 10 == 0:
        wait = WebDriverWait(driver, 10)
        print("break time")
        time.sleep(20)


