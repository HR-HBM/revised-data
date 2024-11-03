import random

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

dataframe = openpyxl.load_workbook("sorted links2.xlsx")
print("workbook loaded")

dataframe1 = dataframe.active
links_list = []
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1):
        links = col[row].value
        links_list.append(links)

def get_random_user_agent():
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36",
        "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:89.0) Gecko/20100101 Firefox/89.0",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/47.0.2526.111 Safari/537.36",
        "Mozilla/5.0 (Linux; Android 10; Pixel 3 XL Build/QP1A.190711.020) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Mobile Safari/537.36",
        "Mozilla/5.0 (iPhone; CPU iPhone OS 14_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15E148 Safari/604.1",
        "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3",
        "Mozilla/5.0 (iPad; CPU OS 14_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15E148 Safari/604.1",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:91.0) Gecko/20100101 Firefox/91.0",
        "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:60.0) Gecko/20100101 Firefox/60.0"
    ]
    return random.choice(user_agents)


chrome_options = webdriver.ChromeOptions()
chrome_options.add_experimental_option("detach", True)

user_agent = get_random_user_agent()
chrome_options.add_argument(f"user-agent={user_agent}")

# # proxy = get_random_proxy()
# chrome_options.add_argument('--proxy-server=http://72.10.160.90:26279')

driver = webdriver.Chrome(options=chrome_options)

maxLink = dataframe1.max_row + 1
currentLinkIndex = 97

while currentLinkIndex <= maxLink:

    name_text = ""
    country_text = ""
    category_text = []
    formatted_categories = ""
    email_text = ""
    number_text = ""
    business_area_text = ""
    website_text = ""

    print(links_list[currentLinkIndex])
    driver.get(f"{links_list[currentLinkIndex]}")


    wait = WebDriverWait(driver, 10)
    item_data = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='finder-profile']/div/div/section/div/div/div[2]/div[5]/button/div/span")))

    try:
        html_list = driver.find_element(By.XPATH, value="//*[@id='finder-products']/div/div/section/div[2]/div/div/div[1]/ul")
        category = html_list.find_elements(By.TAG_NAME, value="li")
        for item in category:
            category_text.append(item.text)
            formatted_categories = ', '.join(category_text)
            print(formatted_categories)
        # category = driver.find_element(By.XPATH, value="//*[@id='profile-title']/h1")
        # name_text = category.text
    except:
        pass

    driver.execute_script("arguments[0].click();", item_data)

    time.sleep(2)



    try:
        name = driver.find_element(By.XPATH, value="//*[@id='profile-title']/h1")
        name_text = name.text
    except:
        pass

    try:
        country = driver.find_element(By.XPATH, value="//*[@id='profile-business-data']/div[1]/div[1]/div/div[2]")
        country_text = country.text
    except:
        pass

    try:
        email = driver.find_element(By.XPATH, value="//*[@id='profile-business-data']/div[1]/div[2]/div[1]/div[1]")
        email_text = email.text
    except:
        pass

    try:
        number = driver.find_element(By.XPATH, value="//*[@id='profile-business-data']/div[1]/div[2]/div[1]/div[2]")
        number_text = number.text
    except:
        pass

    # try:
    #     business_area = driver.find_element(By.XPATH,
    #                                         value="//*[@id='profile-business-data']/div[2]/div/table/tbody/tr[3]/td[2]")
    #     business_area_text = business_area.text
    # except:
    #     pass

    try:
        website = driver.find_element(By.XPATH,
                                      value="//*[@id='profile-business-data']/div[1]/div[2]/div[1]/div[3]/div[2]/ul/li/a")
        website_text = website.text
    except:
        pass

    driver.execute_script("window.open('https://docs.google.com/forms/d/e/1FAIpQLSeABXtU62XVghsKuVzW-aGOazLQE4Rt-Ub8ZOkKbOosAOrTnw/viewform', '_blank');")
    wait = WebDriverWait(driver, 10)


    print("form accessed")
    # Switch to the Google Form tab
    driver.switch_to.window(driver.window_handles[-1])
    time.sleep(4)

    print("form loaded")

    company_name = wait.until(EC.presence_of_element_located((By.XPATH,"//*[@id='mG61Hd']/div[2]/div/div[2]/div[1]/div/div/div[2]/div/div[1]/div/div[1]/input")))
    time.sleep(2)
    company_name.send_keys(name_text)


    company_country = driver.find_element(By.XPATH,value="//*[@id='mG61Hd']/div[2]/div/div[2]/div[2]/div/div/div[2]/div/div[1]/div/div[1]/input")
    company_country.send_keys(country_text)

    company_category = driver.find_element(By.XPATH,value="//*[@id='mG61Hd']/div[2]/div/div[2]/div[3]/div/div/div[2]/div/div[1]/div[2]/textarea")
    company_category.send_keys(formatted_categories)

    company_number = driver.find_element(By.XPATH,value="//*[@id='mG61Hd']/div[2]/div/div[2]/div[4]/div/div/div[2]/div/div[1]/div/div[1]/input")
    company_number.send_keys(number_text)

    company_email = driver.find_element(By.XPATH,value="//*[@id='mG61Hd']/div[2]/div/div[2]/div[5]/div/div/div[2]/div/div[1]/div/div[1]/input")
    company_email.send_keys(email_text)

    company_website = driver.find_element(By.XPATH,value="//*[@id='mG61Hd']/div[2]/div/div[2]/div[6]/div/div/div[2]/div/div[1]/div/div[1]/input")
    company_website.send_keys(website_text)

    submit = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='mG61Hd']/div[2]/div/div[3]/div[1]/div[1]/div")))

    submit.send_keys(Keys.ENTER)

    driver.close()

    driver.switch_to.window(driver.window_handles[-1])
    print(currentLinkIndex)

    currentLinkIndex += 1
    # index_limit = [5, 10, 15, 20, 25]
    if currentLinkIndex % 15 == 0:
        wait = WebDriverWait(driver, 10)
        print("break time")
        time.sleep(20)


